"""Stage 1: parse ``Bank_Eligibility_Matrix.xlsx`` into typed rows.

Reads the ``decision table`` sheet with openpyxl, maps each header to its
snake_case field via :data:`constants.INPUT_COLUMN_TO_FIELD` /
``OUTPUT_COLUMN_TO_FIELD``, and returns one :class:`MatrixRow` per data row. The
cell values are raw ZEN Expression Language strings (e.g. ``">= 675"``,
``"[21..70]"``) — they are *not* evaluated here, only carried forward.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from app.services.rule_engine.constants import (
    INPUT_COLUMN_TO_FIELD,
    MATRIX_SHEET_NAME,
    OUTPUT_COLUMN_TO_FIELD,
)
from app.services.rule_engine.exceptions import MatrixParseError


@dataclass(slots=True)
class MatrixRow:
    """One rule row from the decision table.

    ``inputs`` maps applicant field name -> ZEN condition string (blank cells are
    omitted, meaning "no constraint"). ``outputs`` maps result field name ->
    literal value (e.g. ``bank_name`` -> ``"BOI"``).
    """

    index: int
    inputs: dict[str, str] = field(default_factory=dict)
    outputs: dict[str, str] = field(default_factory=dict)


def _normalize(value: object) -> str:
    """Coerce a cell value to a trimmed string ("" for blank cells)."""
    if value is None:
        return ""
    return str(value).strip()


def parse(matrix_path: Path) -> list[MatrixRow]:
    """Parse the matrix workbook into a list of :class:`MatrixRow`.

    Args:
        matrix_path: Path to the ``.xlsx`` decision table.

    Returns:
        One :class:`MatrixRow` per non-empty data row, in sheet order.

    Raises:
        MatrixParseError: if the sheet is missing, headers are unrecognized, or
            the workbook cannot be opened.
    """
    if not matrix_path.exists():
        raise MatrixParseError(
            f"Matrix workbook not found: {matrix_path}",
            details={"path": str(matrix_path)},
        )

    try:
        workbook = load_workbook(matrix_path, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a variety of IO/zip errors
        raise MatrixParseError(
            f"Could not open matrix workbook: {exc}",
            details={"path": str(matrix_path)},
        ) from exc

    try:
        if MATRIX_SHEET_NAME not in workbook.sheetnames:
            raise MatrixParseError(
                f"Sheet {MATRIX_SHEET_NAME!r} not found in workbook.",
                details={"available": list(workbook.sheetnames)},
            )
        sheet = workbook[MATRIX_SHEET_NAME]

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise MatrixParseError("Matrix sheet is empty.") from None

        headers = [_normalize(cell) for cell in header]
        _check_headers(headers)

        # Build a column-index -> (kind, field) map for the columns we care about.
        column_map: dict[int, tuple[str, str]] = {}
        for col_idx, head in enumerate(headers):
            if head in INPUT_COLUMN_TO_FIELD:
                column_map[col_idx] = ("input", INPUT_COLUMN_TO_FIELD[head])
            elif head in OUTPUT_COLUMN_TO_FIELD:
                column_map[col_idx] = ("output", OUTPUT_COLUMN_TO_FIELD[head])

        parsed: list[MatrixRow] = []
        for row_number, raw in enumerate(rows_iter, start=1):
            cells = [_normalize(cell) for cell in raw]
            if not any(cells):
                continue  # skip fully-blank rows

            inputs: dict[str, str] = {}
            outputs: dict[str, str] = {}
            for col_idx, (kind, field_name) in column_map.items():
                value = cells[col_idx] if col_idx < len(cells) else ""
                if value == "":
                    continue  # blank cell == no constraint / no value
                if kind == "input":
                    inputs[field_name] = value
                else:
                    outputs[field_name] = value

            parsed.append(MatrixRow(index=row_number, inputs=inputs, outputs=outputs))

        return parsed
    finally:
        workbook.close()


def _check_headers(headers: list[str]) -> None:
    """Ensure every required input/output column is present in the header row."""
    present = set(headers)
    missing = [
        col for col in (*INPUT_COLUMN_TO_FIELD, *OUTPUT_COLUMN_TO_FIELD) if col not in present
    ]
    if missing:
        raise MatrixParseError(
            "Matrix is missing required columns.",
            details={"missing": missing, "found": headers},
        )
