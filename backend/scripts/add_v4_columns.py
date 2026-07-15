"""Append the V4 columns to Bank_Eligibility_Matrix.xlsx.

Convention (matches existing matrix):
- boolean 'not allowed' -> 'false'; 'allowed' -> blank (no constraint)
- numeric limits -> ZEN comparison strings
Bank row order in the sheet: BOI, Indian Bank, IOB, BOB, HDFC, AXIS, Kotak, BOM.
"""
from openpyxl import load_workbook

PATH = r"C:\Projects\debtdactory\backend\src\app\services\rule_engine\data\Bank_Eligibility_Matrix.xlsx"

# values ordered [BOI, Indian Bank, IOB, BOB, HDFC, AXIS, Kotak, BOM]
NEW_COLS = [
    ("CIBIL PL Score", [">= 701", "", "", "", "", "", "", ""]),
    ("Age at Last EMI-Salaried", ["<= 60", "<= 60", "<= 75", "<= 60", "<= 70", "<= 70", "<= 70", "<= 70"]),
    ("Age at Last EMI-Self Employed", ["<= 65", "<= 70", "<= 75", "<= 70", "<= 70", "<= 70", "<= 70", "<= 70"]),
    ("Resi-Cum-Office-Owned", ["", "", "", "", "", "", "", ""]),
    ("Resi-Cum-Office-Both Rented", ["", "", "", "false", "", "", "", ""]),
    ("Resi-Office-Separate-Both Rented", ["false", "false", "false", "", "", "", "", "false"]),
    ("Without a Guarantor", ["false", "false", "false", "false", "false", "false", "false", ""]),
    ("With a Guarantor", ["", "", "false", "false", "", "", "", ""]),
    ("Rental Income-Agreement-No ITR-Not in Bank", ["false", "false", "false", "false", "", "", "", "false"]),
    ("Rental Income-Agreement-ITR-Not in Bank", ["", "", "false", "false", "", "", "", "false"]),
    ("Rental Income-Agreement-No ITR-In Bank", ["", "false", "false", "false", "", "", "", "false"]),
    ("SE Current ITR", [">= 300000", ">= 300000", ">= 300000", "", "", "", "", ""]),
    ("SE Previous ITR", ["", ">= 300000", ">= 300000", ">= 300000", ">= 100000", ">= 100000", ">= 100000", ""]),
    ("Business ITR Years", [">= 3", ">= 2", ">= 2", ">= 2", "", "", ">= 2", ">= 2"]),
    ("HUF", ["false", "false", "", "false", "", "", "", "false"]),
    ("Form 16 Years", [">= 2", ">= 2", ">= 2", ">= 1", "", "", "", ">= 2"]),
    ("Co-Applicant Age-Brother", ["", "false", "false", "", "", "", "", "false"]),
    ("Co-Applicant Age-Sister", ["", "false", "false", "", "", "", "", "false"]),
    ("Co-Applicant Income-Brother", ["", "false", "false", "", "", "", "", "false"]),
    ("Co-Applicant Income-Father", ["", "", "", "", "", "", "", ""]),
    ("Co-Applicant Income-Mother", ["", "", "", "", "", "", "", ""]),
    ("Co-Applicant Income-Sister", ["", "false", "false", "false", "", "", "", "false"]),
]

EXPECTED_BANKS = ["BOI", "Indian Bank", "IOB", "BOB", "HDFC", "AXIS", "Kotak", "BOM"]

wb = load_workbook(PATH)
ws = wb["decision table"]

headers = [c.value for c in ws[1]]
bank_col = headers.index("Bank Name") + 1
banks = [ws.cell(row=r, column=bank_col).value for r in range(2, 2 + len(EXPECTED_BANKS))]
assert banks == EXPECTED_BANKS, f"Bank row order changed: {banks}"

existing = set(h for h in headers if h)
next_col = len(headers) + 1
for name, values in NEW_COLS:
    if name in existing:
        print(f"skip (exists): {name}")
        continue
    ws.cell(row=1, column=next_col, value=name)
    for i, v in enumerate(values):
        if v != "":
            ws.cell(row=2 + i, column=next_col, value=v)
    print(f"added col {next_col}: {name}")
    next_col += 1

wb.save(PATH)
print("saved OK")
