import os

import openpyxl


def inspect_xlsx(path):
    print(f"=== Inspecting {path} ===")
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"Sheet '{sheetname}' dimensions: {sheet.dimensions}")
        # Print all rows
        for i, r in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(r is not None for r in r):
                print(f"{i:2d}: {[str(x) if x is not None else '' for x in r]}")
    print("\n")


inspect_xlsx("src/app/services/rule_engine/data/Bank_Eligibility_Matrix.xlsx")
