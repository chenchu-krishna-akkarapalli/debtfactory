import openpyxl


def clean_name(name):
    if not name:
        return ""
    return str(name).strip().replace("\xa0", " ")


# Parameters where "Yes" is a permission (i.e., we accept it, so no constraint)
# and "No" is a restriction (we reject it, so must be false).
PERMISSION_PARAMS = {
    "Rented House-Salaried",
    "Unmarried",
    "Salaried",
    "Employment-Firm",
    "Employment-Pvt Ltd",
    "Employment-Public Ltd",
    "Employment-Govt",
    "Employment-PSU",
    "Rental Income-ITR",
    "Rental Income-reflecting in Bank",
    "Self Employed",
    "Self employed ITR Filled",
    "Propreitorship",
    "Parternship Firm",
    "Private Limited",
    "Public Limited",
}


def translate_value(param, val, bank=None):
    val_str = clean_name(val)
    if not val_str or val_str.upper() in ("NA", "N/A", ""):
        return ""

    param_clean = clean_name(param)

    # Special exact match strings
    if param_clean == "Business Proof":
        if val_str == "Mandatory":
            return '"Mandatory"'
        return f'"{val_str}"'

    # NRI/PIO special case
    if param_clean == "NRI/PIO" and "Condtions" in val_str:
        return ""  # No constraint since they accept it with conditions

    # Special logic for Credit Card Write Off
    if param_clean == "Credit Card Write Off":
        if bank == "BOI":
            return ""  # BOI allows both yes/no
        elif bank == "BOM":
            return "true"  # BOM requires yes
        elif val_str.upper() == "NO":
            return "false"
        return ""

    # Special logic for Existing A/C Holder & Existing Car Loan (requirement params)
    if param_clean in ("Existing A/C Holder", "Existing Car Loan"):
        if val_str.upper() == "YES":
            return "true"
        if val_str.upper() == "NO":
            return ""

    # Boolean fields
    if val_str.upper() == "YES":
        if param_clean in PERMISSION_PARAMS:
            return ""  # "Yes" means accepted/allowed, so no constraint needed
        return "true"

    if val_str.upper() == "NO":
        return "false"  # "No" always means not allowed / must be false

    # Numeric with plus
    if val_str.endswith("+") or val_str.endswith(" +"):
        num_part = val_str.replace("+", "").strip()
        return f">= {num_part}"

    # Numeric stay period or experience
    if "Yrs" in val_str or "Yr" in val_str:
        num_part = val_str.split()[0].strip()
        return f">= {num_part}"

    # Numeric ranges like 21-70
    if "-" in val_str:
        parts = val_str.split("-")
        if len(parts) == 2:
            return f"[{parts[0].strip()}..{parts[1].strip()}]"

    # Write off amount special case
    if "Less than" in val_str:
        num_part = val_str.replace("Less than", "").strip()
        return f"< {num_part}"

    # Numeric bare values (e.g. 25000, 100000)
    if val_str.isdigit():
        if param_clean in ("Minimum Salary", "Min Bus Income"):
            return f">= {val_str}"
        return val_str

    return val_str


def transpose(src_path, dest_path):
    print(f"Transposing {src_path} -> {dest_path}")
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    sheet_src = wb_src.active

    # Read headers (banks) from row 1, starting from column 2 (index 1)
    banks = []
    for col in range(2, sheet_src.max_column + 1):
        cell_val = sheet_src.cell(row=1, column=col).value
        if cell_val:
            banks.append(clean_name(cell_val))

    print("Found banks:", banks)

    # Read parameters and bank values
    rows_data = []
    for r in range(2, sheet_src.max_row + 1):
        param = sheet_src.cell(row=r, column=1).value
        if not param:
            continue
        param = clean_name(param)
        bank_vals = {}
        for idx, bank in enumerate(banks):
            val = sheet_src.cell(row=r, column=idx + 2).value
            bank_vals[bank] = val
        rows_data.append((param, bank_vals))

    # Create new workbook
    wb_dest = openpyxl.Workbook()
    sheet_dest = wb_dest.active
    sheet_dest.title = "decision table"

    # Define headers
    headers = [p[0] for p in rows_data] + ["Bank Name", "Description"]
    sheet_dest.append(headers)

    # Add data row for each bank
    for bank in banks:
        row_cells = []
        for param, bank_vals in rows_data:
            raw_val = bank_vals[bank]
            translated = translate_value(param, raw_val, bank=bank)
            row_cells.append(translated)
        # Add outputs
        row_cells.append(bank)
        row_cells.append("")  # Description
        sheet_dest.append(row_cells)

    wb_dest.save(dest_path)
    print("Transpose completed successfully!")


if __name__ == "__main__":
    transpose(
        "scripts/BRE_sheet.xlsx", "src/app/services/rule_engine/data/Bank_Eligibility_Matrix.xlsx"
    )
    transpose("scripts/BRE_sheet.xlsx", "Bank Eligibility Matrix.xlsx")
